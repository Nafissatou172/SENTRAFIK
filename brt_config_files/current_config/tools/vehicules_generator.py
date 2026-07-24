#!/usr/bin/python
# - - coding: latin-1 - -

import os
import random
import openpyxl
from lxml import etree

def load_vTypes(filename):
	"""Charge les vTypes depuis le fichier XML de routes (brt_road.rou.xml)."""
	vTypes = []
	filepath = os.path.join(os.path.dirname(__file__), filename)
	tree = etree.parse(filepath)
	results = tree.xpath('//routes/vType')
	for r in results:
		vTypes.append(etree.tostring(r, encoding='unicode').strip())
	return vTypes

def get_busStop_names(filename):
	names = []
	filepath = os.path.join(os.path.dirname(__file__), filename)
	tree = etree.parse(filepath)
	results = tree.xpath("//additional/busStop")
	for r in results:
		names.append(r.get('id'))
	return names

def load_busStops(filename, busStop_names):
	vTypes = {}
	busStops = {}
	filepath = os.path.join(os.path.dirname(__file__), filename)
	workbook = openpyxl.load_workbook(filepath)
	sheet = workbook['busStops']
	
	for col in 'CDEFG':
		vTypes[col] = sheet[col + '2'].value
	
	for row in range(4, sheet.max_row+1):
		routes = sheet['A'+str(row)].value.split(' ')
		busStop = sheet['B'+str(row)].value
		if(not busStop in busStop_names):
			continue
		for route in routes:
			if len(route) == 0 :
				continue
		
		if(not route in busStops):
			busStops[route] = {}
			for col in vTypes:
				busStops[route][ vTypes[col] ] = []
		
		for col in 'CDEFG':
			value = sheet[col+str(row)].value
			if((value is None) or (not isinstance(value,str)) or (not ';' in value)):
				continue
			durations = value.split(';')
			busStops[route][ vTypes[col] ].append((busStop, int(durations[0]), int(durations[1])))
	# print(busStops)
	return busStops

def load_routes(filename):
	routes = []
	filepath = os.path.join(os.path.dirname(__file__), filename)
	tree = etree.parse(filepath)
	results = tree.xpath("//routes/route")
	for r in results:
		# print(r.xpath('@id')[0])
		# print(r.get('id'))
		route = '<route edges="' + str(r.get('edges')) + '" color="' + str(r.get('color')) +'" id="' + str(r.get('id')) + '"/>'
		routes.append(route)
	# print(routes)
	return routes

def load_data(filename):
	vTypes = {}
	data = {}
	filepath = os.path.join(os.path.dirname(__file__), filename)
	# see: https://stackoverflow.com/questions/28517508/how-can-i-use-openpyxl-to-read-an-excel-cell-value-and-not-the-formula-computing
	workbook = openpyxl.load_workbook(filepath, data_only=True)
	sheet = workbook['routes']
	
	for col in 'BCDEF':
		vTypes[col] = sheet[col + '2'].value
	
	# for row in range(4, sheet.max_row+1):
	for row in range(4, 20):
		row_data = {}
		for col in 'BCDEF':
			value = sheet[col+str(row)].value
			if((value is None) or (not isinstance(value,int)) or (value<0)):
				value = 0
			row_data[ vTypes[col] ] = int(value)
		route = sheet['A'+str(row)].value
		data[ route ] = row_data
	# print(data)
	return data

def generate(filename, firstDepart = 1, lastDepart = 3600):
	global vTypes
	global routes
	global data
	global busStops
	global nbPietons
	
	# nbVehicules = 0
	# for r in data:
		# for t in data[r]:
			# nbVehicules += data[r][t]q 
	
	# depart_plage = lastDepart - firstDepart
	duree = (lastDepart - firstDepart + 1)
	# durationPlage = maxStopDuration - minStopDuration
	id = 0
	vehicules = []
	for r in data:
		# nbVehicules = 0
		# for t in data[r]:
			# nbVehicules += data[r][t]
		# pas_des_departs = max(1, duree / nbVehicules)
		
		for t in data[r]:
			nbVehicules = data[r][t]
			if(nbVehicules == 0):
				continue
			nbBusStops = 0
			if((r in busStops) and (t in busStops[r])):
				nbBusStops = len(busStops[r][t])
			pas_des_departs = max(1, duree / nbVehicules)
			for k in range(data[r][t]):
				id += 1
				depart = firstDepart + ((k * pas_des_departs) % duree)
				# see: https://sumo.dlr.de/docs/Definition_of_Vehicles,_Vehicle_Types,_and_Routes.html
				# xml = '\n<vehicle jmCrossingGap="20" id="v' + str(id) + '" type="' + t + '" route="' + r + '" depart="' + str(depart) + '" departSpeed="speedLimit"'
				# xml = '\n<vehicle id="v' + str(id) + '" type="' + t + '" route="' + r + '" depart="' + str(depart) + '" departSpeed="speedLimit"'
				xml = '\n<vehicle id="v' + str(id) + '" type="' + t + '" route="' + r + '" depart="' + str(depart) + '" departSpeed="max" departLane="free" arrivalLane="current" '
				if(t == 'Particulier'):
					xml += ' personNumber="' + str(random.randint(1, 2)) +'"/>'
				elif(t == 'Taxi'):
					if(nbBusStops > 0):
						xml += ' personNumber="' + str(random.randint(1, 3)) +'">'
						if(random.random() < 0.4):
							pos = min(nbBusStops-1, random.randint(0, 1))
							# duration = random.randint(3, 10)
							duration = random.randint(busStops[r][t][pos][1], busStops[r][t][pos][2])
							xml += '\n\t<stop busStop="' + busStops[r][t][pos][0] + '" duration="' + str(duration) + '" parking="true"/>'
						xml += "\n</vehicle>"
					else:
						xml += "/>"
				elif(t == 'DakarDemDikk'):
					if(nbBusStops > 0):
						xml += ' personNumber="' + str(random.randint(70, 82)) +'">'
						if(random.random() < 0.85):
							pos = min(nbBusStops-1, random.randint(0, 1))
							# duration = minStopDuration + random.randint(0, durationPlage)
							# duration = 5 + random.randint(0, 10)
							duration = random.randint(busStops[r][t][pos][1], busStops[r][t][pos][2])
							xml += '\n\t<stop busStop="' + busStops[r][t][pos][0] + '" duration="' + str(duration) + '" parking="true"/>'
							if(pos == 0 and random.random() < 0.75):
								# duration = minStopDuration + random.randint(0, durationPlage)
								duration = random.randint(busStops[r][t][pos][1], busStops[r][t][pos][2])
								xml += '\n\t<stop busStop="' + busStops[r][t][0][0] + '" duration="' + str(duration) + '" parking="true"/>'
						xml += "\n</vehicle>"
					else:
						xml += "/>"
				else: # CarRapide, NdiagaNdiaye, TATA
					if(nbBusStops > 0):
						if(t == 'TATA'):
							xml += ' personNumber="' + str(random.randint(38, 42)) +'">'
						else:
							xml += ' personNumber="' + str(random.randint(24, 28)) +'">'
						if(random.random() < 0.75):
							pos = min(nbBusStops-1, random.randint(0, 1))
							# duration = minStopDuration + random.randint(0, durationPlage)
							duration = random.randint(busStops[r][t][pos][1], busStops[r][t][pos][2])
							xml += '\n\t<stop busStop="' + busStops[r][t][pos][0] + '" duration="' + str(duration) + '" parking="false"/>'
						xml += "\n</vehicle>"
					else:
						xml += "/>"
				vehicules.append({'xml':xml, 'depart':depart})
	# Tri des véhicules par ordre d'entrée dans la simulation
	for i in range(len(vehicules)-1):
		for j in range(i+1, len(vehicules)):
			if(vehicules[i]['depart'] > vehicules[j]['depart']):
				tmp = vehicules[i]
				vehicules[i] = vehicules[j]
				vehicules[j] = tmp
	
	filepath = os.path.join(os.path.dirname(__file__), filename)
	file = open(filepath, 'w')
	file.write('<routes xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="http://sumo.dlr.de/xsd/routes_file.xsd">\n\n')
	for vType in vTypes:
		file.write(vType + "\n")
	file.write("\n")
	for route in routes:
		file.write(route + "\n")
	
	# file.write('\n<personFlow id="p" begin="0" end="' + str(lastDepart) + '" number="' + str(nbPietons) + '" probability="0.5">\n')
	file.write('\n<personFlow id="p1" begin="0" end="' + str(lastDepart) + '" probability="0.1">\n')
	# file.write('\n<personFlow id="p" begin="0" number="' + str(nbPietons) + '" probability="0.5">\n')
	file.write('\t<walk route="passagePietons_NS_ok"/>\n')
	file.write('</personFlow>\n')
	
	file.write('\n<personFlow id="p2" begin="0" end="' + str(lastDepart) + '" probability="0.1">\n')
	file.write('\t<walk route="passagePietons_NS_ko"/>\n')
	file.write('</personFlow>\n')
	
	for v in vehicules:
		file.write(v['xml'])
	file.write('\n</routes>')
	file.close()

filename = 'current_state.rou.xml'
firstDepart = 1
lastDepart = 3600
nbPietons = 200
vTypes = load_vTypes('../config/brt_road.rou.xml')
routes = load_routes('../config/brt_road.rou.xml')
busStop_names = get_busStop_names('../config/output.add.xml')
busStops = load_busStops('data_distribution_congestion.xlsx', busStop_names)
data = load_data('data_distribution_congestion.xlsx')
generate(filename, firstDepart, lastDepart)

print(' --> All tasks ended!')