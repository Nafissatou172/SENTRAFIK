#!/usr/bin/env python3
"""
scenario2.py — Scénario "Policier Intelligent" pour le projet SENTRAFIK.
Ce script implémente une gestion dynamique des feux de circulation (TLS)
en fonction de la demande en temps réel (files d'attente) avec priorité absolue au BRT.
"""

import os
import sys
import argparse
import time
from collections import defaultdict
from datetime import datetime

# ─────────────────────────────────────────────────────────────
# Configuration SUMO
# ─────────────────────────────────────────────────────────────
if 'SUMO_HOME' in os.environ:
    os.environ['PROJ_LIB'] = '/Library/Frameworks/EclipseSUMO.framework/Versions/1.25.0/EclipseSUMO/framework/EclipseSUMO.framework/Versions/1.25.0/EclipseSUMO/share/proj'
    tools = os.path.join(os.environ['SUMO_HOME'], 'tools')
    sys.path.append(tools)
else:
    sys.exit("Erreur: Veuillez déclarer la variable d'environnement 'SUMO_HOME'")

import traci

# ─────────────────────────────────────────────────────────────
# Constantes
# ─────────────────────────────────────────────────────────────
SUMO_BINARY_GUI = '/Library/Frameworks/EclipseSUMO.framework/Versions/Current/EclipseSUMO/bin/sumo-gui'
SUMO_BINARY_CMD = '/Library/Frameworks/EclipseSUMO.framework/Versions/Current/EclipseSUMO/bin/sumo'
SUMO_CONFIG = 'brt.sumocfg'
STEP_LENGTH = 1
BRT_FLOW_PREFIXES = ['brt_flow_b1', 'brt_flow_b2', 'brt_flow_b3']
SAMPLING_INTERVAL = 10  # Pour la collecte des métriques

# ─────────────────────────────────────────────────────────────
# Contrôleur "Policier V2" (Préemption & Sécurité)
# ─────────────────────────────────────────────────────────────
class PolicemanController:
    """Version améliorée avec détection anticipée du BRT et gestion du jaune."""

    def __init__(self, brt_prefixes, min_green_time=15, preemption_dist=100):
        self.brt_prefixes = brt_prefixes
        self.min_green_time = min_green_time
        self.preemption_dist = preemption_dist
        
        self.tls_states = {} # {tls_id: {'last_switch': time, 'target_phase': idx, 'status': 'NORMAL'|'TRANSITION'}}
        self.tls_program = {} # {tls_id: [phases]}
        self.tls_lane_map = {} # {tls_id: {phase_idx: [lanes]}}

    def initialize(self):
        """Initialise la connaissance du réseau et des programmes de feux."""
        for tls_id in traci.trafficlight.getIDList():
            try:
                # Récupération de la logique
                logic = traci.trafficlight.getCompleteRedYellowGreenDefinition(tls_id)[0]
                # getControlledLinks() est plus précis pour mapper les indices de phase
                links = traci.trafficlight.getControlledLinks(tls_id)
                
                self.tls_program[tls_id] = logic.phases
                self.tls_states[tls_id] = {
                    'last_switch': -self.min_green_time,
                    'status': 'NORMAL',
                    'target_phase': traci.trafficlight.getPhase(tls_id)
                }
                
                # Mapping phase_index -> liste des voies qui reçoivent le vert
                self.tls_lane_map[tls_id] = {}
                for p_idx, phase in enumerate(logic.phases):
                    if 'G' in phase.state or 'g' in phase.state:
                        green_lanes = []
                        for i, s in enumerate(phase.state):
                            if s in 'Gg' and i < len(links):
                                # links[i] est une liste de tuples (fromLane, toLane, viaLane)
                                for link in links[i]:
                                    if link:
                                        green_lanes.append(link[0]) # La voie d'entrée
                        self.tls_lane_map[tls_id][p_idx] = list(set(green_lanes))
            except Exception as e:
                print(f"⚠️  Erreur initialisation TLS {tls_id}: {e}")

    def get_brt_priorities(self):
        """Détecte les BRT s'approchant des feux et retourne les TLS impactés."""
        priorities = defaultdict(set) # {tls_id: {target_phase_indices}}
        
        # Parcourir tous les véhicules pour trouver les BRT
        for veh_id in traci.vehicle.getIDList():
            if any(veh_id.startswith(pref) for pref in self.brt_prefixes):
                # getNextTLS retourne [(tlsID, tlsIndex, distance, state), ...]
                next_tls = traci.vehicle.getNextTLS(veh_id)
                if next_tls:
                    tls_id, link_index, dist, state = next_tls[0]
                    if dist < self.preemption_dist:
                        # --- NOUVELLE LOGIQUE DE SÉCURITÉ ---
                        speed = traci.vehicle.getSpeed(veh_id)
                        is_stopped_at_station = traci.vehicle.getStopState(veh_id) & 1
                        
                        # Si le bus est à l'arrêt ou quasi-immobile loin du feu, on n'active pas la priorité
                        if is_stopped_at_station or (speed < 0.2 and dist > 25):
                            continue
                            
                        # Trouver quelle phase donne le vert à cet index de lien
                        program = self.tls_program.get(tls_id, [])
                        for p_idx, phase in enumerate(program):
                            # Vérifier si l'index du lien est au vert dans cette phase
                            if link_index < len(phase.state):
                                if phase.state[link_index] in 'Gg':
                                    priorities[tls_id].add(p_idx)
        return priorities

    def control_step(self, sim_time):
        """Logique de décision et de transition."""
        brt_priorities = self.get_brt_priorities()
        
        for tls_id in self.tls_states:
            state = self.tls_states[tls_id]
            current_phase = traci.trafficlight.getPhase(tls_id)
            
            # 1. Gérer les transitions en cours (ne pas interrompre le orange)
            current_state_str = traci.trafficlight.getRedYellowGreenState(tls_id)
            if 'y' in current_state_str or 'u' in current_state_str:
                state['status'] = 'TRANSITION'
                continue
            else:
                state['status'] = 'NORMAL'

            # 2. Vérifier le temps de vert minimum (sauf si priorité BRT critique ?)
            if sim_time - state['last_switch'] < self.min_green_time:
                # On ne change pas si on vient de changer, SAUF si un BRT arrive
                # mais on garde une stabilité minimale de 5s quand même
                if tls_id not in brt_priorities or (sim_time - state['last_switch'] < 5):
                    continue

            # 3. Calculer la meilleure phase
            best_phase = current_phase
            max_score = -1
            
            # Phases candidates (uniquement celles avec du vert)
            candidate_phases = self.tls_lane_map.get(tls_id, {})
            
            for p_idx, lanes in candidate_phases.items():
                score = 0
                # Heuristique : File d'attente (poids 1) + Temps d'attente cumulé (poids 0.1)
                # Heuristique : File d'attente (poids 2.0) + Temps d'attente cumulé (poids 0.2)
                for lane in lanes:
                    score += traci.lane.getLastStepHaltingNumber(lane) * 2.0
                    score += traci.lane.getWaitingTime(lane) * 0.2
                
                # Priorité BRT (Préemption intelligente)
                if tls_id in brt_priorities and p_idx in brt_priorities[tls_id]:
                    score += 400 # Priorité équilibrée
                
                if score > max_score:
                    max_score = score
                    best_phase = p_idx

            # 4. Exécuter le changement
            if best_phase != current_phase:
                # Transition sécurisée : passer par le orange si défini
                # Dans SUMO, en static, la phase i+1 est souvent le orange de la phase i
                next_p = (current_phase + 1) % len(self.tls_program[tls_id])
                next_state = self.tls_program[tls_id][next_p].state
                
                if 'y' in next_state or 'u' in next_state:
                    # On passe par le orange intermédiaire
                    traci.trafficlight.setPhase(tls_id, next_p)
                else:
                    # Pas de orange détecté, switch direct (rare mais possible)
                    traci.trafficlight.setPhase(tls_id, best_phase)
                
                state['last_switch'] = sim_time
                state['target_phase'] = best_phase

# ─────────────────────────────────────────────────────────────
# Collecteur de Métriques (Adapté de scenario1.py)
# ─────────────────────────────────────────────────────────────
class TrafficMetricsCollector:
    def __init__(self):
        self.step_count = 0
        self.sim_time = 0.0
        
        # Métriques par véhicule BRT (accumulées)
        self.brt_data = {}  # {veh_id: {metrics...}}

        # Métriques des feux de circulation
        self.tls_data = defaultdict(lambda: {
            'total_waiting_time': 0.0,
            'total_queue_length': 0,
            'total_vehicles_passed': 0,
            'sample_count': 0,
            'controlled_lanes': [],
        })

        # Suivi précis des métriques par véhicule pour le résumé global
        self.veh_final_waiting = {}  # {veh_id: cumulative_waiting}
        self.veh_final_loss = {}     # {veh_id: cumulative_loss}

        # Suivi des véhicules par feux pour compter les passages
        self.tls_seen_vehicles = defaultdict(set)

        # Métriques globales cumulées
        self.global_metrics = {
            'total_vehicles_departed': 0,
            'total_vehicles_arrived': 0,
            'total_teleports': 0,
            'total_halting_vehicles': 0,
            'cumulative_speed': 0.0,
            'speed_sample_count': 0,
        }
        self.seen_vehicles = set()

    def is_brt_vehicle(self, veh_id):
        return any(veh_id.startswith(prefix) for prefix in BRT_FLOW_PREFIXES)

    def collect_step_metrics(self):
        self.step_count += 1
        self.sim_time = traci.simulation.getTime()

        # 1. Compteurs globaux (À CHAQUE PAS)
        self.global_metrics['total_vehicles_departed'] += traci.simulation.getDepartedNumber()
        self.global_metrics['total_vehicles_arrived'] += traci.simulation.getArrivedNumber()
        self.global_metrics['total_teleports'] += traci.simulation.getStartingTeleportNumber()

        # 2. Détecter les arrivées BRT et départs (À CHAQUE PAS pour précision)
        for arrived_id in traci.simulation.getArrivedIDList():
            if self.is_brt_vehicle(arrived_id):
                if arrived_id in self.brt_data:
                    self.brt_data[arrived_id]['arrival_time'] = self.sim_time
        
        # Détection des départs BRT pour un depart_time précis
        for departed_id in traci.simulation.getDepartedIDList():
            if self.is_brt_vehicle(departed_id):
                if departed_id not in self.brt_data:
                    self.brt_data[departed_id] = {
                        'id': departed_id,
                        'depart_time': self.sim_time,
                        'arrival_time': None,
                        'total_waiting_time': 0.0,
                        'total_time_loss': 0.0,
                        'distance': 0.0,
                        'speeds': [],
                        'energy_consumed': 0.0,
                        'co2': 0.0,
                        'fuel': 0.0,
                        'route_id': traci.vehicle.getRouteID(departed_id),
                        'type': traci.vehicle.getTypeID(departed_id),
                    }

        # 3. Métriques échantillonnées (Toutes les SAMPLING_INTERVAL secondes)
        if self.step_count % SAMPLING_INTERVAL == 0:
            vehicle_ids = traci.vehicle.getIDList()
            num_vehicles = len(vehicle_ids)
            total_speed_step = 0.0
            total_halting = 0

            for veh_id in vehicle_ids:
                speed = traci.vehicle.getSpeed(veh_id)
                waiting_time = traci.vehicle.getAccumulatedWaitingTime(veh_id)
                time_loss = traci.vehicle.getTimeLoss(veh_id)
                co2 = traci.vehicle.getCO2Emission(veh_id)
                fuel = traci.vehicle.getFuelConsumption(veh_id)
                
                total_speed_step += speed
                if speed < 0.1:
                    total_halting += 1

                # Suivi global
                self.seen_vehicles.add(veh_id)
                self.veh_final_loss[veh_id] = time_loss
                self.veh_final_waiting[veh_id] = waiting_time

                # ── Métriques BRT spécifiques ──
                if self.is_brt_vehicle(veh_id):
                    # Initialisation si pas déjà fait par getDepartedIDList
                    if veh_id not in self.brt_data:
                        self.brt_data[veh_id] = {
                            'id': veh_id,
                            'depart_time': self.sim_time,
                            'arrival_time': None,
                            'total_waiting_time': 0.0,
                            'total_time_loss': 0.0,
                            'distance': 0.0,
                            'speeds': [],
                            'energy_consumed': 0.0,
                            'co2': 0.0,
                            'fuel': 0.0,
                            'route_id': traci.vehicle.getRouteID(veh_id),
                            'type': traci.vehicle.getTypeID(veh_id),
                        }

                    brt = self.brt_data[veh_id]
                    if brt['arrival_time'] is None:
                        brt['total_waiting_time'] = waiting_time
                        brt['total_time_loss'] = time_loss
                        brt['distance'] = traci.vehicle.getDistance(veh_id)
                        brt['speeds'].append(speed)
                        
                        interval = SAMPLING_INTERVAL
                        brt['co2'] += (co2 / 1000.0) * interval
                        brt['fuel'] += fuel * interval
                        try:
                            brt['energy_consumed'] += (traci.vehicle.getElectricityConsumption(veh_id) / 3600.0) * interval
                        except: pass

            if num_vehicles > 0:
                self.global_metrics['cumulative_speed'] += total_speed_step / num_vehicles
                self.global_metrics['speed_sample_count'] += 1
            self.global_metrics['total_halting_vehicles'] += total_halting
            
            # 4. Feux de circulation
            tls_ids = traci.trafficlight.getIDList()
            for tls_id in tls_ids:
                controlled_lanes = set(traci.trafficlight.getControlledLanes(tls_id))
                queue_length = 0
                current_vehs = set()
                for lane in controlled_lanes:
                    queue_length += traci.lane.getLastStepHaltingNumber(lane)
                    current_vehs.update(traci.lane.getLastStepVehicleIDs(lane))

                passed_count = 0
                if tls_id in self.tls_seen_vehicles:
                    passed_count = len(self.tls_seen_vehicles[tls_id] - current_vehs)
                
                tls = self.tls_data[tls_id]
                tls['total_waiting_time'] += queue_length * SAMPLING_INTERVAL 
                tls['total_queue_length'] += queue_length
                tls['total_vehicles_passed'] += passed_count
                tls['sample_count'] += 1
                self.tls_seen_vehicles[tls_id] = current_vehs
                if not tls['controlled_lanes']:
                    tls['controlled_lanes'] = list(controlled_lanes)

            

    def finalize_brt_data(self):
        for data in self.brt_data.values():
            s = data.pop('speeds', [])
            data['vitesse_moyenne_km_h'] = round((sum(s)/len(s) if s else 0)*3.6, 2)
            data['nb_echantillons'] = len(s)
            data['distance_km'] = round(data['distance']/1000.0, 3)
            # Durée réelle du trajet (Arrivée - Départ)
            arrival = data['arrival_time'] if data['arrival_time'] else self.sim_time
            duration = arrival - data['depart_time']
            data['duree_dans_simulation_s'] = round(duration, 2)
            data['total_waiting_time'] = round(data['total_waiting_time'], 2)
            data['total_time_loss'] = round(data['total_time_loss'], 2)
            data['energy_consumed_wh'] = round(data['energy_consumed'], 4)
            data['co2_g'] = round(data['co2'], 4)
            data['fuel_ml'] = round(data['fuel'], 4)

    def get_global_summary(self):
        gm = self.global_metrics
        avg_speed = (gm['cumulative_speed'] / gm['speed_sample_count'] if gm['speed_sample_count'] > 0 else 0)
        n_veh = max(gm['total_vehicles_departed'], 1)
        
        total_loss = sum(self.veh_final_loss.values())
        total_waiting = sum(self.veh_final_waiting.values())

        return {
            'Date de simulation': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Durée simulée (min)': round(self.sim_time / 60, 2),
            'Pas de simulation': self.step_count,
            'Véhicules déployés': gm['total_vehicles_departed'],
            'Véhicules arrivés': gm['total_vehicles_arrived'],
            'Taux complétion (%)': round(gm['total_vehicles_arrived'] / n_veh * 100, 2),
            'Vitesse moyenne réseau (km/h)': round(avg_speed * 3.6, 2),
            'Temps attente moyen/veh (s)': round(total_waiting / n_veh, 2),
            'Temps perdu en moyenne/veh (min)': round((total_loss / n_veh) / 60, 2) if n_veh > 0 else 0,
            'Téléportations (congestion)': gm['total_teleports'],
            'Nb Total bus BRT suivis': len(self.brt_data),
            'Nb feux de circulation': len(self.tls_data),
        }

    def get_tls_summary(self):
        res = []
        for tid, d in self.tls_data.items():
            n = max(d['sample_count'], 1)
            passed = max(d['total_vehicles_passed'], 1)
            res.append({
                'id_feu': tid, 
                'temps_attente_moy_par_veh_min': round((d['total_waiting_time'] / passed) / 60, 2),
                'file_attente_moyenne': round(d['total_queue_length']/n, 2),
                'vehicules_passes': d['total_vehicles_passed'], 
                'scenario': 'Policier'
            })
        res.sort(key=lambda x: x['temps_attente_moy_par_veh_min'], reverse=True)
        return res

# ─────────────────────────────────────────────────────────────
# Utils: Export Excel
# ─────────────────────────────────────────────────────────────
def export_to_excel(collector, output_file):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Erreur: openpyxl requis.")
        return

    wb = openpyxl.Workbook()
    
    # ── Styles ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_fill_green = PatternFill(start_color="2D936C", end_color="2D936C", fill_type="solid")
    header_fill_purple = PatternFill(start_color="7B2D8E", end_color="7B2D8E", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, fill=header_fill):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[col_letter].width = min(max_length + 4, 35)

    # ══════════════════════════════════════════════════════════
    # Feuille 1 : Résumé Global
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = "Résumé Global"
    ws1.append(["Métrique", "Valeur"])
    for k, v in collector.get_global_summary().items(): 
        ws1.append([k, v])
    style_header(ws1, header_fill)
    auto_width(ws1)

    # ══════════════════════════════════════════════════════════
    # Feuille 2 : BRT Véhicules
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("BRT Véhicules")
    headers = [
        'ID Véhicule', 'Type', 'Route', 'Temps Départ (s)', 'Durée du trajet (h)', 
        'Distance (km)', 'Vitesse Moy (km/h)', 'Temps Attente Total (s)', 'Temps Total perdu (h)', 
        'Énergie Consommée (Wh)', 'CO2 (g)', 'Carburant (ml)'
    ]
    ws2.append(headers)
    collector.finalize_brt_data()
    for d in sorted(collector.brt_data.values(), key=lambda x: x['id']):
        ws2.append([
            d['id'], d['type'], d['route_id'], d['depart_time'], 
            round(d['duree_dans_simulation_s'] / 3600, 2), 
            d['distance_km'], d['vitesse_moyenne_km_h'], 
            d['total_waiting_time'], round(d['total_time_loss'] / 3600, 4),
            d['energy_consumed_wh'], d['co2_g'], d['fuel_ml']
        ])
    style_header(ws2, header_fill_green)
    auto_width(ws2)

    # ══════════════════════════════════════════════════════════
    # Feuille 3 : Feux de Circulation
    # ══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Feux de Circulation")
    ws4.append(['ID Feu', 'Temps Attente Moyen/veh (min)', 'File Attente Moyenne', 'Véhicules Passés', 'Scénario'])
    for r in collector.get_tls_summary(): 
        ws4.append(list(r.values()))
    style_header(ws4, header_fill_purple)
    auto_width(ws4)

    os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
    wb.save(output_file)
    print(f"✅ Résultats exportés : {output_file}")

def final_reporting(collector, start_time, args):
    """Affiche le résumé final dans la console."""
    elapsed_total = time.time() - start_time
    summary = collector.get_global_summary()
    tls_summary = collector.get_tls_summary()
    
    print("\n" + "=" * 60)
    print("  RÉSULTATS")
    print("=" * 60)
    print(f"  ⏱️  Durée simulée     : {summary['Durée simulée (min)']} min")
    print(f"  🚗 Véhicules déployés : {summary['Véhicules déployés']}")
    print(f"  ✅ Véhicules arrivés  : {summary['Véhicules arrivés']}")
    print(f"  📊 Taux complétion    : {summary['Taux complétion (%)']}%")
    print(f"  🏎️  Vitesse moyenne    : {summary['Vitesse moyenne réseau (km/h)']} km/h")
    print(f"  ⚠️  Téléportations     : {summary['Téléportations (congestion)']}")
    print(f"  🚌 Bus BRT suivis     : {summary['Nb Total bus BRT suivis']}")
    print(f"  🚦 Feux analysés      : {summary['Nb feux de circulation']}")
    print(f"  ⏱️  Temps réel         : {elapsed_total:.1f}s")
    print("=" * 60)

    # ── Export Excel ──
    export_to_excel(collector, args.output)
    
    print(f"\n✅ Résultats exportés dans : {args.output}")
    print(f"   📊 {len(summary)} métriques globales")
    print(f"   🚌 {len(collector.brt_data)} véhicules BRT suivis")
    print(f"   🚦 {len(tls_summary)} feux de circulation analysés")
    print("\n🎉 Terminé ! Ouvrez le fichier Excel pour analyser les résultats.")


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="SENTRAFIK Scenario 2: Contrôle Policier Intelligent (Dakar BRT)")
    parser.add_argument('--no-gui', action='store_true',
                        help="Lancer SUMO sans l'interface graphique (mode console)")
    parser.add_argument('--output', '-o', default='data/scenario2_results.xlsx',
                        help="Nom du fichier Excel de sortie (défaut: data/scenario2_results.xlsx)")
    parser.add_argument('--min-green', type=int, default=10, 
                        help="Temps de vert minimum (s)")
    parser.add_argument('--step-length', type=float, default=STEP_LENGTH,
                        help=f"Durée d'un pas de simulation (défaut: {STEP_LENGTH})")
    args = parser.parse_args()

    # Gestion du dossier data/
    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, 'data')
    os.makedirs(data_dir, exist_ok=True)
    output_basename = os.path.basename(args.output)
    args.output = os.path.join(data_dir, output_basename)

    sumo_binary = SUMO_BINARY_CMD if args.no_gui else SUMO_BINARY_GUI
    sumo_cmd = [
        sumo_binary,
        '-c', SUMO_CONFIG,
        '--step-length', str(args.step_length),
        '--no-warnings', 'true'
    ]

    if not args.no_gui:
        sumo_cmd.extend(['--delay', '0'])

    print("=" * 60)
    print("  SENTRAFIK — Collecte de Métriques de Trafic")
    print("  BRT Dakar — Scénario 2: Policier Intelligent (IA)")
    print("=" * 60)
    print(f"\n🔧 Mode : {'Headless (rapide)' if args.no_gui else 'GUI'}")
    print(f"📁 Config : {SUMO_CONFIG}")
    print(f"📊 Sortie : {args.output}")
    print(f"⏱️  Step length : {args.step_length}s")
    print(f"📐 Échantillonnage : toutes les {SAMPLING_INTERVAL * args.step_length}s simulées")

    # ── Lancement ──
    print("\n⏳ Lancement de la simulation SUMO...")
    start_time = time.time()
    traci.start(sumo_cmd)
    
    collector = TrafficMetricsCollector()
    controller = PolicemanController(BRT_FLOW_PREFIXES, min_green_time=args.min_green)
    controller.initialize()

    print("🔄 Simulation en cours...")
    last_progress = -1

    try:
        while traci.simulation.getMinExpectedNumber() > 0:
            traci.simulationStep()
            sim_time = traci.simulation.getTime()
            
            # Contrôle dynamique des feux
            controller.control_step(sim_time)
            
            # Collecte des métriques
            collector.collect_step_metrics()
            
            # Progression console toutes les 5 minutes simulées
            progress = int(sim_time / 60)
            if progress > last_progress and progress % 5 == 0:
                elapsed = time.time() - start_time
                arr = collector.global_metrics['total_vehicles_arrived']
                print(f"   ⏱️  {progress} min simulées | {arr} véhicules arrivés | {elapsed:.0f}s réelles")
                last_progress = progress
                
    except KeyboardInterrupt:
        print("\n⚠️  Simulation interrompue par l'utilisateur.")
    except Exception as e:
        print(f"\n⚠️  Erreur TraCI: {e}")
    finally:
        try:
            traci.close()
        except:
            pass

    # Affichage du résumé final
    final_reporting(collector, start_time, args)


if __name__ == '__main__':
    main()
