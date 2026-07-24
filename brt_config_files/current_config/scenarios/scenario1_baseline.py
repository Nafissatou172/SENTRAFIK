#!/usr/bin/env python3
"""
traffic_metrics.py — Collecte de métriques de trafic SUMO pour le projet SENTRAFIK.

Ce script lance une simulation SUMO, collecte les métriques d'évaluation du trafic
via TraCI, et exporte les résultats dans un fichier Excel pour comparaison
avant/après optimisation IA des feux de circulation.

Usage:
    python traffic_metrics.py                  # Mode GUI par défaut (comme stats.py)
    python traffic_metrics.py --no-gui         # Mode headless (rapide, sans interface)
    python traffic_metrics.py --output mon_fichier.xlsx  # Nom de fichier personnalisé
"""

import os
import sys
import argparse
import time
from collections import defaultdict
from datetime import datetime

# ─────────────────────────────────────────────────────────────
# Configuration SUMO
# ─────────────────────────────────────────────────────────────
if 'SUMO_HOME' in os.environ:
    os.environ['PROJ_LIB'] = '/Library/Frameworks/EclipseSUMO.framework/Versions/1.25.0/EclipseSUMO/framework/EclipseSUMO.framework/Versions/1.25.0/EclipseSUMO/share/proj'
    tools = os.path.join(os.environ['SUMO_HOME'], 'tools')
    sys.path.append(tools)
else:
    sys.exit("Erreur: Veuillez déclarer la variable d'environnement 'SUMO_HOME'")

import traci

# ─────────────────────────────────────────────────────────────
# Constantes
# ─────────────────────────────────────────────────────────────
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
SUMO_BINARY_GUI = '/Library/Frameworks/EclipseSUMO.framework/Versions/Current/EclipseSUMO/bin/sumo-gui'
SUMO_BINARY_CMD = '/Library/Frameworks/EclipseSUMO.framework/Versions/Current/EclipseSUMO/bin/sumo'
SUMO_CONFIG = os.path.join(PROJECT_ROOT, 'config', 'brt.sumocfg')
STEP_LENGTH = 1  # Durée d'un pas de simulation en secondes (Valeur idéale pour RL)

# IDs des flux BRT à surveiller (préfixes)
BRT_FLOW_PREFIXES = ['brt_flow_b1', 'brt_flow_b2', 'brt_flow_b3']

# Intervalle d'échantillonnage pour les séries temporelles (en pas de simulation)
# 50 pas * 0.1s = 5 secondes de temps simulé
#Par pas de 5s
SAMPLING_INTERVAL = 10


# ─────────────────────────────────────────────────────────────
# Classe principale de collecte de métriques
# ─────────────────────────────────────────────────────────────
class TrafficMetricsCollector:
    """Collecte et agrège les métriques de trafic pendant une simulation SUMO."""

    def __init__(self):
        # Compteurs globaux
        self.step_count = 0
        self.sim_time = 0.0

        # Métriques par véhicule BRT (accumulées)
        self.brt_data = {}  # {veh_id: {metrics...}}

        # Métriques des feux de circulation
        self.tls_data = defaultdict(lambda: {
            'total_waiting_time': 0.0,
            'total_queue_length': 0,
            'total_vehicles_passed': 0,
            'sample_count': 0,
            'controlled_lanes': [],
        })

        # Suivi précis des métriques par véhicule pour le résumé global
        self.veh_final_waiting = {}  # {veh_id: cumulative_waiting}
        self.veh_final_loss = {}     # {veh_id: cumulative_loss}

        # Suivi des véhicules par feux pour compter les passages
        self.tls_seen_vehicles = defaultdict(set)

        # Métriques globales cumulées
        self.global_metrics = {
            'total_waiting_time': 0.0,
            'total_time_loss': 0.0,
            'total_vehicles_departed': 0,
            'total_vehicles_arrived': 0,
            'total_teleports': 0,
            'total_halting_vehicles': 0, # vehicules à l'arrets 
            'cumulative_speed': 0.0,
            'speed_sample_count': 0,
            #'total_stops': 0,
        }

        # Suivi des véhicules déjà vus (pour compter les arrivées)
        self.seen_vehicles = set()
        self.arrived_vehicles = set()

    def is_brt_vehicle(self, veh_id):
        """Vérifie si un véhicule est un bus BRT."""
        return any(veh_id.startswith(prefix) for prefix in BRT_FLOW_PREFIXES)

    def collect_step_metrics(self):
        """Collecte les métriques de simulation. Les métriques lourdes sont échantillonnées."""
        self.step_count += 1
        self.sim_time = traci.simulation.getTime()

        # 1. Compteurs globaux (À CHAQUE PAS)
        self.global_metrics['total_vehicles_departed'] += traci.simulation.getDepartedNumber()
        self.global_metrics['total_vehicles_arrived'] += traci.simulation.getArrivedNumber()
        self.global_metrics['total_teleports'] += traci.simulation.getStartingTeleportNumber()

        # 2. Détecter les arrivées BRT (À CHAQUE PAS)
        for arrived_id in traci.simulation.getArrivedIDList():
            if arrived_id in self.brt_data:
                self.brt_data[arrived_id]['arrival_time'] = self.sim_time

        # 3. Métriques échantillonnées (Toutes les SAMPLING_INTERVAL secondes)
        if self.step_count % SAMPLING_INTERVAL == 0:
            vehicle_ids = traci.vehicle.getIDList()
            num_vehicles = len(vehicle_ids)
            total_speed_step = 0.0
            total_halting = 0

            for veh_id in vehicle_ids:
                speed = traci.vehicle.getSpeed(veh_id)
                waiting_time = traci.vehicle.getAccumulatedWaitingTime(veh_id)
                time_loss = traci.vehicle.getTimeLoss(veh_id)
                
                total_speed_step += speed
                if speed < 0.1:
                    total_halting += 1

                # Suivi global
                self.seen_vehicles.add(veh_id)
                self.veh_final_loss[veh_id] = time_loss
                self.veh_final_waiting[veh_id] = waiting_time

                # Métriques BRT
                if any(veh_id.startswith(p) for p in BRT_FLOW_PREFIXES):
                    if veh_id not in self.brt_data:
                        self.brt_data[veh_id] = {
                            'id': veh_id,
                            'type': traci.vehicle.getTypeID(veh_id),
                            'route_id': traci.vehicle.getRouteID(veh_id),
                            'depart_time': self.sim_time,
                            'arrival_time': None,
                            'speeds': [],
                            'distance': 0.0,
                            'total_waiting_time': 0.0,
                            'total_time_loss': 0.0,
                            'energy_consumed': 0.0,
                            'co2': 0.0,
                            'fuel': 0.0
                        }
                    
                    brt = self.brt_data[veh_id]
                    if brt['arrival_time'] is None:
                        brt['speeds'].append(speed)
                        brt['distance'] = traci.vehicle.getDistance(veh_id)
                        brt['total_waiting_time'] = waiting_time
                        brt['total_time_loss'] = time_loss
                        
                        # Accumulation des consommations
                        interval = SAMPLING_INTERVAL
                        brt['co2'] += (traci.vehicle.getCO2Emission(veh_id) / 1000.0) * interval
                        brt['fuel'] += traci.vehicle.getFuelConsumption(veh_id) * interval
                        try:
                            brt['energy_consumed'] += (traci.vehicle.getElectricityConsumption(veh_id) / 3600.0) * interval
                        except: pass

            if num_vehicles > 0:
                self.global_metrics['cumulative_speed'] += total_speed_step / num_vehicles
                self.global_metrics['speed_sample_count'] += 1
            self.global_metrics['total_halting_vehicles'] += total_halting

            # 4. Feux de circulation
            tls_ids = traci.trafficlight.getIDList()
            for tls_id in tls_ids:
                controlled_lanes = set(traci.trafficlight.getControlledLanes(tls_id))
                queue_length = 0
                current_vehs = set()
                for lane in controlled_lanes:
                    queue_length += traci.lane.getLastStepHaltingNumber(lane)
                    current_vehs.update(traci.lane.getLastStepVehicleIDs(lane))

                passed_count = 0
                if tls_id in self.tls_seen_vehicles:
                    passed_count = len(self.tls_seen_vehicles[tls_id] - current_vehs)
                
                tls = self.tls_data[tls_id]
                tls['total_waiting_time'] += queue_length * SAMPLING_INTERVAL
                tls['total_queue_length'] += queue_length
                tls['total_vehicles_passed'] += passed_count
                tls['sample_count'] += 1
                self.tls_seen_vehicles[tls_id] = current_vehs
                if not tls['controlled_lanes']:
                    tls['controlled_lanes'] = list(controlled_lanes)
                tls['sample_count'] += 1
                
                # Mettre à jour la liste des véhicules vus pour le prochain échantillon
                self.tls_seen_vehicles[tls_id] = current_vehs
                
                if not tls['controlled_lanes']:
                    tls['controlled_lanes'] = list(controlled_lanes)


    def finalize_brt_data(self):
        """Finalise les données BRT en calculant les moyennes."""
        for veh_id, data in self.brt_data.items():
            speeds = data.pop('speeds', [])
            data['vitesse_moyenne_m_s'] = round(sum(speeds) / len(speeds), 2) if speeds else 0
            data['vitesse_moyenne_km_h'] = round(data['vitesse_moyenne_m_s'] * 3.6, 2)
            data['nb_echantillons'] = len(speeds)
            data['distance_km'] = round(data['distance'] / 1000.0, 3)
            # Durée réelle du trajet (Arrivée - Départ)
            arrival = data['arrival_time'] if data['arrival_time'] else self.sim_time
            duration = arrival - data['depart_time']
            data['duree_dans_simulation_s'] = round(duration, 2)
            data['total_waiting_time'] = round(data['total_waiting_time'], 2)
            data['total_time_loss'] = round(data['total_time_loss'], 2)
            #data['max_waiting_time'] = round(data['max_waiting_time'], 2)
            data['energy_consumed_wh'] = round(data['energy_consumed'], 4)
            data['co2_g'] = round(data['co2'], 4)
            data['fuel_ml'] = round(data['fuel'], 4)

    def get_global_summary(self):
        """Retourne le résumé global des métriques."""
        gm = self.global_metrics
        avg_speed = (gm['cumulative_speed'] / gm['speed_sample_count']
                     if gm['speed_sample_count'] > 0 else 0)
        
        # Moyenne par véhicule (basé sur ceux qui ont démarré)
        n_veh = max(gm['total_vehicles_departed'], 1)
        
        # Le total est la somme des dernières valeurs cumulées connues pour chaque véhicule
        total_loss = sum(self.veh_final_loss.values())
        total_waiting = sum(self.veh_final_waiting.values())

        return {
            'Date de simulation': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            #'Durée simulée (s)': round(self.sim_time, 2),
            'Durée simulée (min)': round(self.sim_time / 60, 2),
            'Pas de simulation': self.step_count,
            'Véhicules déployés': gm['total_vehicles_departed'],
            'Véhicules arrivés (throughput)': gm['total_vehicles_arrived'],
            'Taux de complétion (%)': round(gm['total_vehicles_arrived'] / n_veh * 100, 2),
            'Vitesse moyenne réseau (km/h)': round(avg_speed * 3.6, 2),
            'Temps attente moyen/veh (s)': round(total_waiting / n_veh, 2),
            'Temps perdu en moyenne/veh (min)': round((total_loss / n_veh) / 60, 2) if n_veh > 0 else 0,
            'Téléportations (congestion)': gm['total_teleports'],
            'Nb total véhicules BRT suivis': len(self.brt_data),
            'Nb feux de circulation': len(self.tls_data),
        }

    def get_tls_summary(self):
        """Retourne les métriques par feu de circulation."""
        results = []
        for tls_id, data in self.tls_data.items():
            n = max(data['sample_count'], 1)
            passed = max(data['total_vehicles_passed'], 1)
            results.append({
                'id_feu': tls_id,
                'temps_attente_moy_par_veh_min': round((data['total_waiting_time'] / passed) / 60, 2),
                'file_attente_moyenne': round(data['total_queue_length'] / n, 2),
                'vehicules_passes': data['total_vehicles_passed'],
                'scenario': 'Baseline (Temps Fixe)',
            })
        # Trier par retard moyen décroissant
        results.sort(key=lambda x: x['temps_attente_moy_par_veh_min'], reverse=True)
        return results


# ─────────────────────────────────────────────────────────────
# Export vers Excel
# ─────────────────────────────────────────────────────────────
def export_to_excel(collector, output_file):
    """Exporte toutes les métriques collectées vers un fichier Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("Erreur: openpyxl est requis. Installez-le avec: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Styles ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_fill_green = PatternFill(start_color="2D936C", end_color="2D936C", fill_type="solid")
    header_fill_orange = PatternFill(start_color="E88D2A", end_color="E88D2A", fill_type="solid")
    header_fill_purple = PatternFill(start_color="7B2D8E", end_color="7B2D8E", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, fill=header_fill):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[col_letter].width = min(max_length + 4, 35)

    # ══════════════════════════════════════════════════════════
    # Feuille 1 : Résumé Global
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Résumé Global"
    summary = collector.get_global_summary()
    ws1.append(["Métrique", "Valeur"])
    for key, value in summary.items():
        ws1.append([key, value])
    style_header(ws1, header_fill)
    auto_width(ws1)

    # ══════════════════════════════════════════════════════════
    # Feuille 2 : BRT Véhicules
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("BRT Véhicules")
    brt_headers = [
        'ID Véhicule', 'Type', 'Route',
        'Temps Départ (s)', 'Durée du trajet (h)',
        'Distance (km)', 'Vitesse Moy (km/h)',
        'Temps Attente Total (s)', 'Temps Total Perdu (h)',
        'Énergie Consommée (Wh)', 'CO2 (g)', 'Carburant (ml)'
    ]
    ws2.append(brt_headers)

    collector.finalize_brt_data()
    for veh_id in sorted(collector.brt_data.keys()):
        d = collector.brt_data[veh_id]
        ws2.append([
            d['id'], d['type'], d['route_id'],
            d['depart_time'], round(d['duree_dans_simulation_s'] / 3600, 2),
            d['distance_km'], d['vitesse_moyenne_km_h'],
            d['total_waiting_time'], round(d['total_time_loss'] / 3600, 4),
            d['energy_consumed_wh'], d['co2_g'], d['fuel_ml']
        ])
    style_header(ws2, header_fill_green)
    auto_width(ws2)


    # ══════════════════════════════════════════════════════════
    # Feuille 3 : Feux de Circulation
    # ══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Feux de Circulation")
    tls_headers = [
        'ID Feu','Temps Attente Moyen/veh (min)',
        'File Attente Moyenne', 'Véhicules Passés', 'Scénario'
    ]
    ws4.append(tls_headers)
    tls_summary = collector.get_tls_summary()
    for row in tls_summary:
        ws4.append(list(row.values()))
    style_header(ws4, header_fill_purple)
    auto_width(ws4)

    # ── Sauvegarder ──
    os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
    wb.save(output_file)
    print(f"\n✅ Résultats exportés dans : {output_file}")
    print(f"   📊 {len(summary)} métriques globales")
    print(f"   🚌 {len(collector.brt_data)} véhicules BRT suivis")
    #print(f"   📈 {len(collector.timeseries)} points de série temporelle")
    print(f"   🚦 {len(tls_summary)} feux de circulation analysés")


# ─────────────────────────────────────────────────────────────
# Programme principal
# ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Collecte de métriques de trafic SUMO pour le projet SENTRAFIK (BRT Dakar)")
    parser.add_argument('--no-gui', action='store_true',
                        help="Lancer SUMO sans l'interface graphique (mode console)")
    parser.add_argument('--output', '-o', default='data/scenario1/traffic_metrics_results.xlsx',
                        help="Nom du fichier Excel de sortie (défaut: data/scenario1/traffic_metrics_results.xlsx)")
    parser.add_argument('--step-length', type=float, default=STEP_LENGTH,
                        help=f"Durée d'un pas de simulation en secondes (défaut: {STEP_LENGTH})")
    args = parser.parse_args()

    # Forcer l'enregistrement dans le dossier data/ (relatif au projet)
    data_dir = os.path.join(PROJECT_ROOT, 'data', 'scenario1')
    os.makedirs(data_dir, exist_ok=True)
    # Si l'utilisateur passe juste un nom de fichier, on le place dans data/scenario1/
    output_basename = os.path.basename(args.output)
    args.output = os.path.join(data_dir, output_basename)

    # Choix du binaire SUMO (GUI par défaut comme dans stats.py)
    sumo_binary = SUMO_BINARY_CMD if args.no_gui else SUMO_BINARY_GUI

    sumo_cmd = [
        sumo_binary,
        '-c', SUMO_CONFIG,
        '--step-length', str(args.step_length),
        '--no-warnings', 'true'
        #'--lateral-resolution', '0.1'  # Mis à jour pour un réalisme accru pavé par 0.8m
    ]

    if not args.no_gui:
        sumo_cmd.extend(['--delay', '0']) # Délai par défaut comme dans stats.py

    print("=" * 60)
    print("  SENTRAFIK — Collecte de Métriques de Trafic")
    print("  BRT Dakar — Évaluation avant/après optimisation IA")
    print("=" * 60)
    print(f"\n🔧 Mode : {'Headless (rapide)' if args.no_gui else 'GUI'}")
    print(f"📁 Config : {SUMO_CONFIG}")
    print(f"📊 Sortie : {args.output}")
    print(f"⏱️  Step length : {args.step_length}s")
    print(f"📐 Échantillonnage : toutes les {SAMPLING_INTERVAL * args.step_length}s simulées")

    # ── Lancement de la simulation ──
    print("\n⏳ Lancement de la simulation SUMO...")
    start_time = time.time()
    traci.start(sumo_cmd)

    collector = TrafficMetricsCollector()

    print("🔄 Simulation en cours...")

    # Barre de progression simple
    last_progress = -1

    try:
        while traci.simulation.getMinExpectedNumber() > 0:
            traci.simulationStep()
            collector.collect_step_metrics()

            # Afficher la progression toutes les 60 secondes simulées
            sim_time = collector.sim_time
            progress = int(sim_time / 60)
            if progress > last_progress and progress % 5 == 0:
                elapsed = time.time() - start_time
                print(f"   ⏱️  {progress} min simulées | "
                      f"{collector.global_metrics['total_vehicles_arrived']} véhicules arrivés | "
                      f"{elapsed:.0f}s réelles écoulées")
                last_progress = progress

    except KeyboardInterrupt:
        print("\n⚠️  Simulation interrompue par l'utilisateur. Export des données collectées...")
    except traci.exceptions.FatalTraCIError:
        print("\n⚠️  Connexion SUMO fermée. Export des données collectées...")
    finally:
        try:
            traci.close()
        except Exception:
            pass

    elapsed_total = time.time() - start_time

    # ── Résumé console ──
    summary = collector.get_global_summary()
    print("\n" + "=" * 60)
    print("  RÉSULTATS")
    print("=" * 60)
    print(f"  ⏱️  Durée simulée     : {summary['Durée simulée (min)']} min")
    print(f"  🚗 Véhicules déployés : {summary['Véhicules déployés']}")
    print(f"  ✅ Véhicules arrivés  : {summary['Véhicules arrivés (throughput)']}")
    print(f"  📊 Taux complétion    : {summary['Taux de complétion (%)']}%")
    print(f"  🏎️  Vitesse moyenne    : {summary['Vitesse moyenne réseau (km/h)']} km/h")
    print(f"  ⚠️  Téléportations     : {summary['Téléportations (congestion)']}")
    print(f"  🚌 Bus BRT suivis     : {summary['Nb total véhicules BRT suivis']}")
    print(f"  🚦 Feux analysés      : {summary['Nb feux de circulation']}")
    print(f"  ⏱️  Temps réel         : {elapsed_total:.1f}s")
    print("=" * 60)

    # ── Export Excel ──
    export_to_excel(collector, args.output)

    print("\n🎉 Terminé ! Ouvrez le fichier Excel pour analyser les résultats.")


if __name__ == '__main__':
    main()
