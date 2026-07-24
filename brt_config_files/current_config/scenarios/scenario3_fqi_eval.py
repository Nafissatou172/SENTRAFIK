#!/usr/bin/env python3
"""
fqi_eval.py — Phase 3 : Évaluation des modèles FQI entraînés
==============================================================
Lance la simulation SUMO en utilisant les modèles Random Forest
entraînés par fqi_train.py pour contrôler les feux. Collecte les
métriques de performance et exporte un fichier Excel comparable
aux scénarios 1 (baseline) et 2 (policier).

Usage :
    python fqi_eval.py [--models data/fqi_models/] [--output data/scenario3_fqi.xlsx]
"""

import os
import sys
import math
import pickle
import logging
import argparse
import time
import numpy as np
from datetime import datetime
from collections import defaultdict
from typing import Dict, List
import xml.etree.ElementTree as ET

# ── SUMO / TraCI ──────────────────────────────────────────────────────────────
if 'SUMO_HOME' in os.environ:
    os.environ['PROJ_LIB'] = '/Library/Frameworks/EclipseSUMO.framework/Versions/1.25.0/EclipseSUMO/framework/EclipseSUMO.framework/Versions/1.25.0/EclipseSUMO/share/proj'
    tools = os.path.join(os.environ['SUMO_HOME'], 'tools')
    sys.path.append(tools)
else:
    sys.exit("Erreur: Veuillez déclarer la variable d'environnement 'SUMO_HOME'")

import traci

# ═════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═════════════════════════════════════════════════════════════════════════════

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
SUMO_BINARY = "/Library/Frameworks/EclipseSUMO.framework/Versions/Current/EclipseSUMO/bin/sumo"
SUMO_CFG    = os.path.join(PROJECT_ROOT, 'config', 'brt.sumocfg')
NET_XML     = os.path.join(PROJECT_ROOT, 'config', 'output.net.xml')

STEP_LENGTH       = 1
MIN_GREEN_STEPS   = 15
NEIGHBOR_RADIUS   = 600
MAX_NEIGHBORS     = 4
DECISION_INTERVAL = 5      # décision tous les N pas (Version 1)
BRT_PREFIXES      = ['brt_flow_b1', 'brt_flow_b2', 'brt_flow_b3']
SAMPLING_INTERVAL = 10     # échantillonnage métriques

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger(__name__)


# ═════════════════════════════════════════════════════════════════════════════
#  TOPOLOGIE (identique à fqi_collector.py)
# ═════════════════════════════════════════════════════════════════════════════

def load_topology(net_xml_path: str):
    tree = ET.parse(net_xml_path)
    root = tree.getroot()
    junctions = {}
    for junc in root.findall("junction"):
        if "traffic_light" in junc.get("type", ""):
            jid = junc.get("id")
            junctions[jid] = {
                "x": float(junc.get("x", 0)),
                "y": float(junc.get("y", 0)),
                "incLanes": junc.get("incLanes", "").split(),
            }
    num_phases = {}
    for tl in root.findall("tlLogic"):
        tlid = tl.get("id")
        num_phases[tlid] = len(tl.findall("phase"))
    valid_ids = set(num_phases.keys())
    junctions = {jid: info for jid, info in junctions.items() if jid in valid_ids}
    neighbors = {jid: set() for jid in valid_ids}
    for edge in root.findall("edge"):
        if edge.get("function") in ("internal", "walkingarea"):
            continue
        frm, to = edge.get("from"), edge.get("to")
        if frm in valid_ids and to in valid_ids and frm != to:
            neighbors[frm].add(to)
            neighbors[to].add(frm)
    tl_list = list(valid_ids)
    for i, a in enumerate(tl_list):
        ax, ay = junctions[a]["x"], junctions[a]["y"]
        for b in tl_list[i + 1:]:
            bx, by = junctions[b]["x"], junctions[b]["y"]
            if math.hypot(ax - bx, ay - by) <= NEIGHBOR_RADIUS:
                neighbors[a].add(b)
                neighbors[b].add(a)
    for jid in valid_ids:
        neighbors[jid].discard(jid)
        neighbors[jid] = sorted(
            list(neighbors[jid]),
            key=lambda nb: math.hypot(
                junctions[jid]["x"] - junctions[nb]["x"],
                junctions[jid]["y"] - junctions[nb]["y"],
            ),
        )[:MAX_NEIGHBORS]
    agents_info = {}
    for jid in valid_ids:
        agents_info[jid] = {
            "x": junctions[jid]["x"], "y": junctions[jid]["y"],
            "incLanes": junctions[jid]["incLanes"],
            "n_phases": num_phases[jid],
            "neighbors": neighbors.get(jid, []),
        }
    return agents_info


# ═════════════════════════════════════════════════════════════════════════════
#  FONCTIONS UTILITAIRES (identiques à fqi_collector.py)
# ═════════════════════════════════════════════════════════════════════════════

def get_queues(inc_lanes):
    queues = []
    for lane in inc_lanes:
        try:
            queues.append(float(traci.lane.getLastStepHaltingNumber(lane)))
        except traci.exceptions.TraCIException:
            queues.append(0.0)
    return queues


def build_state(tls_id, info, shared_states):
    queues = get_queues(info["incLanes"])
    try:
        phase = traci.trafficlight.getPhase(tls_id)
    except traci.exceptions.TraCIException:
        phase = 0
    n_phases = info["n_phases"]
    vec = list(queues)
    vec.append(phase / max(n_phases - 1, 1))
    vec.append(n_phases / 10.0)
    brt_present = 0.0
    for lane in info["incLanes"]:
        try:
            for veh_id in traci.lane.getLastStepVehicleIDs(lane):
                if any(veh_id.startswith(p) for p in BRT_PREFIXES):
                    brt_present = 1.0
                    break
        except traci.exceptions.TraCIException:
            pass
        if brt_present > 0:
            break
    vec.append(brt_present)
    for nb_id in info["neighbors"]:
        if nb_id in shared_states:
            ns = shared_states[nb_id]
            vec.append(ns["queue_mean"])
            vec.append(ns["phase_norm"])
        else:
            vec.extend([0.0, 0.0])
    padded = len(info["neighbors"])
    while padded < MAX_NEIGHBORS:
        vec.extend([0.0, 0.0])
        padded += 1
    return np.array(vec, dtype=np.float32), queues, phase


def apply_action(tls_id, action, n_phases, last_switch, step):
    if action == 0 or step - last_switch < MIN_GREEN_STEPS:
        return last_switch
    try:
        current = traci.trafficlight.getPhase(tls_id)
        next_phase = (current + 1) % n_phases
        traci.trafficlight.setPhase(tls_id, next_phase)
        return step
    except traci.exceptions.TraCIException:
        return last_switch


# ═════════════════════════════════════════════════════════════════════════════
#  COLLECTEUR DE MÉTRIQUES (adapté de scenario2.py)
# ═════════════════════════════════════════════════════════════════════════════

class TrafficMetricsCollector:
    def __init__(self):
        self.step_count = 0
        self.sim_time = 0.0
        self.brt_data = {}
        self.tls_data = defaultdict(lambda: {
            'total_waiting_time': 0.0, 'total_queue_length': 0,
            'total_vehicles_passed': 0, 'sample_count': 0,
            'controlled_lanes': [],
        })
        self.veh_final_waiting = {}
        self.veh_final_loss = {}
        self.tls_seen_vehicles = defaultdict(set)
        self.global_metrics = {
            'total_vehicles_departed': 0, 'total_vehicles_arrived': 0,
            'total_teleports': 0, 'total_halting_vehicles': 0,
            'cumulative_speed': 0.0, 'speed_sample_count': 0,
        }
        self.seen_vehicles = set()

    def is_brt(self, veh_id):
        return any(veh_id.startswith(p) for p in BRT_PREFIXES)

    def collect(self):
        self.step_count += 1
        self.sim_time = traci.simulation.getTime()
        gm = self.global_metrics
        gm['total_vehicles_departed'] += traci.simulation.getDepartedNumber()
        gm['total_vehicles_arrived'] += traci.simulation.getArrivedNumber()
        gm['total_teleports'] += traci.simulation.getStartingTeleportNumber()

        # BRT arrivées
        for aid in traci.simulation.getArrivedIDList():
            if self.is_brt(aid) and aid in self.brt_data:
                self.brt_data[aid]['arrival_time'] = self.sim_time

        # BRT départs
        for did in traci.simulation.getDepartedIDList():
            if self.is_brt(did) and did not in self.brt_data:
                self.brt_data[did] = {
                    'id': did, 'depart_time': self.sim_time,
                    'arrival_time': None, 'total_waiting_time': 0.0,
                    'total_time_loss': 0.0, 'distance': 0.0, 'speeds': [],
                    'energy_consumed': 0.0, 'co2': 0.0, 'fuel': 0.0,
                    'route_id': traci.vehicle.getRouteID(did),
                    'type': traci.vehicle.getTypeID(did),
                }

        # Échantillonnage
        if self.step_count % SAMPLING_INTERVAL == 0:
            vids = traci.vehicle.getIDList()
            n = len(vids)
            total_speed = 0.0
            halting = 0
            for vid in vids:
                speed = traci.vehicle.getSpeed(vid)
                wt = traci.vehicle.getAccumulatedWaitingTime(vid)
                tl = traci.vehicle.getTimeLoss(vid)
                total_speed += speed
                if speed < 0.1:
                    halting += 1
                self.seen_vehicles.add(vid)
                self.veh_final_loss[vid] = tl
                self.veh_final_waiting[vid] = wt
                if self.is_brt(vid):
                    if vid not in self.brt_data:
                        self.brt_data[vid] = {
                            'id': vid, 'depart_time': self.sim_time,
                            'arrival_time': None, 'total_waiting_time': 0.0,
                            'total_time_loss': 0.0, 'distance': 0.0, 'speeds': [],
                            'energy_consumed': 0.0, 'co2': 0.0, 'fuel': 0.0,
                            'route_id': traci.vehicle.getRouteID(vid),
                            'type': traci.vehicle.getTypeID(vid),
                        }
                    brt = self.brt_data[vid]
                    if brt['arrival_time'] is None:
                        brt['total_waiting_time'] = wt
                        brt['total_time_loss'] = tl
                        brt['distance'] = traci.vehicle.getDistance(vid)
                        brt['speeds'].append(speed)
                        brt['co2'] += (traci.vehicle.getCO2Emission(vid) / 1000.0) * SAMPLING_INTERVAL
                        brt['fuel'] += traci.vehicle.getFuelConsumption(vid) * SAMPLING_INTERVAL
                        try:
                            brt['energy_consumed'] += (traci.vehicle.getElectricityConsumption(vid) / 3600.0) * SAMPLING_INTERVAL
                        except:
                            pass
            if n > 0:
                gm['cumulative_speed'] += total_speed / n
                gm['speed_sample_count'] += 1
            gm['total_halting_vehicles'] += halting

            # Feux
            for tid in traci.trafficlight.getIDList():
                lanes = set(traci.trafficlight.getControlledLanes(tid))
                ql = sum(traci.lane.getLastStepHaltingNumber(l) for l in lanes)
                vehs = set()
                for l in lanes:
                    vehs.update(traci.lane.getLastStepVehicleIDs(l))
                passed = len(self.tls_seen_vehicles[tid] - vehs) if tid in self.tls_seen_vehicles else 0
                d = self.tls_data[tid]
                d['total_waiting_time'] += ql * SAMPLING_INTERVAL
                d['total_queue_length'] += ql
                d['total_vehicles_passed'] += passed
                d['sample_count'] += 1
                self.tls_seen_vehicles[tid] = vehs
                if not d['controlled_lanes']:
                    d['controlled_lanes'] = list(lanes)

    def finalize_brt(self):
        for data in self.brt_data.values():
            s = data.pop('speeds', [])
            data['vitesse_moyenne_km_h'] = round((sum(s)/len(s) if s else 0)*3.6, 2)
            data['distance_km'] = round(data['distance']/1000.0, 3)
            arrival = data['arrival_time'] if data['arrival_time'] else self.sim_time
            data['duree_s'] = round(arrival - data['depart_time'], 2)
            data['energy_wh'] = round(data['energy_consumed'], 4)
            data['co2_g'] = round(data['co2'], 4)
            data['fuel_ml'] = round(data['fuel'], 4)

    def summary(self):
        gm = self.global_metrics
        avg_speed = (gm['cumulative_speed'] / gm['speed_sample_count'] if gm['speed_sample_count'] > 0 else 0)
        n = max(gm['total_vehicles_departed'], 1)
        total_loss = sum(self.veh_final_loss.values())
        total_waiting = sum(self.veh_final_waiting.values())
        return {
            'Date de simulation': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Durée simulée (min)': round(self.sim_time / 60, 2),
            'Pas de simulation': self.step_count,
            'Véhicules déployés': gm['total_vehicles_departed'],
            'Véhicules arrivés': gm['total_vehicles_arrived'],
            'Taux complétion (%)': round(gm['total_vehicles_arrived'] / n * 100, 2),
            'Vitesse moyenne réseau (km/h)': round(avg_speed * 3.6, 2),
            'Temps attente moyen/veh (s)': round(total_waiting / n, 2),
            'Temps perdu en moyenne/veh (min)': round((total_loss / n) / 60, 2),
            'Téléportations (congestion)': gm['total_teleports'],
            'Nb Total bus BRT suivis': len(self.brt_data),
            'Nb feux de circulation': len(self.tls_data),
            'Scénario': 'FQI (Random Forest)',
        }

    def tls_summary(self):
        res = []
        for tid, d in self.tls_data.items():
            n = max(d['sample_count'], 1)
            passed = max(d['total_vehicles_passed'], 1)
            res.append({
                'id_feu': tid,
                'temps_attente_moy_min': round((d['total_waiting_time'] / passed) / 60, 2),
                'file_attente_moy': round(d['total_queue_length'] / n, 2),
                'vehicules_passes': d['total_vehicles_passed'],
                'scenario': 'FQI-RF',
            })
        res.sort(key=lambda x: x['temps_attente_moy_min'], reverse=True)
        return res


# ═════════════════════════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ═════════════════════════════════════════════════════════════════════════════

def export_excel(collector, output_file):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl requis. Installez-le avec : pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    hfill_g = PatternFill(start_color="2D936C", end_color="2D936C", fill_type="solid")
    hfill_p = PatternFill(start_color="7B2D8E", end_color="7B2D8E", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    def style_h(ws, fill):
        for c in ws[1]:
            c.font = hf; c.fill = fill
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            c.border = border

    def auto_w(ws):
        for col in ws.columns:
            ml = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 35)
            for c in col:
                c.border = border
                if c.row > 1:
                    c.alignment = Alignment(horizontal='center')

    # Résumé
    ws1 = wb.active; ws1.title = "Résumé Global"
    ws1.append(["Métrique", "Valeur"])
    for k, v in collector.summary().items():
        ws1.append([k, v])
    style_h(ws1, hfill); auto_w(ws1)

    # BRT
    ws2 = wb.create_sheet("BRT Véhicules")
    ws2.append(['ID', 'Type', 'Route', 'Départ (s)', 'Durée (h)',
                'Distance (km)', 'Vitesse Moy (km/h)',
                'Attente (s)', 'Temps Perdu (h)',
                'Énergie (Wh)', 'CO2 (g)', 'Carburant (ml)'])
    collector.finalize_brt()
    for d in sorted(collector.brt_data.values(), key=lambda x: x['id']):
        ws2.append([
            d['id'], d['type'], d['route_id'], d['depart_time'],
            round(d['duree_s'] / 3600, 2), d['distance_km'],
            d['vitesse_moyenne_km_h'],
            round(d['total_waiting_time'], 2), round(d['total_time_loss'] / 3600, 4),
            d['energy_wh'], d['co2_g'], d['fuel_ml'],
        ])
    style_h(ws2, hfill_g); auto_w(ws2)

    # Feux
    ws3 = wb.create_sheet("Feux de Circulation")
    ws3.append(['ID Feu', 'Attente Moy/veh (min)', 'File Moy', 'Véhicules Passés', 'Scénario'])
    for r in collector.tls_summary():
        ws3.append(list(r.values()))
    style_h(ws3, hfill_p); auto_w(ws3)

    os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
    wb.save(output_file)
    log.info(f"✅ Résultats exportés : {output_file}")


# ═════════════════════════════════════════════════════════════════════════════
#  BOUCLE D'ÉVALUATION
# ═════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="FQI Eval — Évaluation des modèles entraînés")
    parser.add_argument('--models', '-m', default='data/fqi_models/',
                        help="Dossier des modèles entraînés")
    parser.add_argument('--output', '-o', default='data/scenario3/scenario3_fqi_results.xlsx',
                        help="Fichier Excel de sortie")
    parser.add_argument('--no-gui', action='store_true',
                        help="Désactiver le GUI SUMO")
    args = parser.parse_args()

    args.models = os.path.join(PROJECT_ROOT, args.models)
    args.output = os.path.join(PROJECT_ROOT, args.output)
    os.makedirs(os.path.dirname(args.output) or '.', exist_ok=True)

    # Charger topologie et modèles
    agents_info = load_topology(NET_XML)
    models = {}
    model_count = 0
    for tls_id in agents_info:
        model_path = os.path.join(args.models, f"{tls_id}.pkl")
        if os.path.exists(model_path):
            with open(model_path, 'rb') as f:
                models[tls_id] = pickle.load(f)
                model_count += 1

    log.info(f"Modèles chargés : {model_count}/{len(agents_info)}")

    # Lancer SUMO
    binary = SUMO_BINARY if args.no_gui else SUMO_BINARY.replace('/sumo', '/sumo-gui')
    sumo_cmd = [binary, "-c", SUMO_CFG, "--step-length", str(STEP_LENGTH), "--no-warnings", "true"]
    traci.start(sumo_cmd)

    collector = TrafficMetricsCollector()
    last_switches = {tls_id: -MIN_GREEN_STEPS for tls_id in agents_info}
    start_time = time.time()
    step = 0

    print("=" * 60)
    print("  SENTRAFIK — Scénario 3 : FQI (Random Forest)")
    print("=" * 60)

    try:
        while traci.simulation.getMinExpectedNumber() > 0:
            # Décision IA tous les DECISION_INTERVAL pas
            if step % DECISION_INTERVAL == 0:
                shared = {}
                for tls_id, info in agents_info.items():
                    qs = get_queues(info["incLanes"])
                    try:
                        ph = traci.trafficlight.getPhase(tls_id)
                    except:
                        ph = 0
                    shared[tls_id] = {
                        "queue_mean": float(np.mean(qs)) if qs else 0.0,
                        "phase_norm": ph / max(info["n_phases"] - 1, 1),
                    }

                for tls_id, info in agents_info.items():
                    state_vec, qs, ph = build_state(tls_id, info, shared)

                    # Choisir l'action via le modèle RF
                    if tls_id in models and step - last_switches[tls_id] >= MIN_GREEN_STEPS:
                        model = models[tls_id]
                        # Batch : 2 actions en un seul predict()
                        X_batch = np.array([
                            np.append(state_vec, 0.0),
                            np.append(state_vec, 1.0),
                        ])
                        q_values = model.predict(X_batch)
                        action = int(np.argmax(q_values))
                    else:
                        action = 0  # Pas de modèle → garder la phase

                    last_switches[tls_id] = apply_action(
                        tls_id, action, info["n_phases"],
                        last_switches[tls_id], step
                    )

            traci.simulationStep()
            collector.collect()
            step += 1

            # Progress
            progress = int(traci.simulation.getTime() / 60)
            if step % 300 == 0:
                elapsed = time.time() - start_time
                arr = collector.global_metrics['total_vehicles_arrived']
                print(f"   ⏱️  {progress} min simulées | {arr} véhicules arrivés | {elapsed:.0f}s réelles")

    except KeyboardInterrupt:
        log.info("Interruption manuelle.")
    finally:
        traci.close()

    # Résumé
    elapsed = time.time() - start_time
    s = collector.summary()
    print(f"\n{'='*60}")
    print(f"  RÉSULTATS")
    print(f"{'='*60}")
    print(f"  ⏱️  Durée simulée     : {s['Durée simulée (min)']} min")
    print(f"  🚗 Véhicules déployés : {s['Véhicules déployés']}")
    print(f"  ✅ Véhicules arrivés  : {s['Véhicules arrivés']}")
    print(f"  📊 Taux complétion    : {s['Taux complétion (%)']}%")
    print(f"  🏎️  Vitesse moyenne    : {s['Vitesse moyenne réseau (km/h)']} km/h")
    print(f"  ⚠️  Téléportations     : {s['Téléportations (congestion)']}")
    print(f"  🚌 Bus BRT suivis     : {s['Nb Total bus BRT suivis']}")
    print(f"  ⏱️  Temps réel         : {elapsed:.0f}s")
    print(f"{'='*60}")

    export_excel(collector, args.output)


if __name__ == "__main__":
    main()
